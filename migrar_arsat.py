
from KMZReader.kmz_reader import LectorKMZ
from KMZReader.db_migration import MigrarInfo
from openpyxl import load_workbook

# Suponiendo que ya tienes el contenido_kml del archivo KML
PROVINCIA = 'MENDOZA'
ruta_kmz = f'C:/Users/user/Desktop/Lambda/migraciones/MIGRACION ARSAT/Mendoza/{PROVINCIA}.kmz'
path_excel_tramo = 'C:/Users/user/Desktop/Lambda/migraciones/MIGRACION ARSAT/240104 ATRIBUTOS TRAMOS Y DERIVACIONES KMZ V68.xlsx'
path_excel_sitios = 'C:/Users/user/Desktop/Lambda/migraciones/MIGRACION ARSAT/240104 SITIOS - (Neuquen) - ACTUALIZACION KMZV68.xlsx'
lector = LectorKMZ()
# En el objeto MigrarInfo se asignan las variables del ID de la compania y las variables que seran cargadas tambien en el sidx
migrador = MigrarInfo(company_id='30',
                    variables_sidx=['@oName',
                                    '@acronimo',
                                    '@nombre',
                                    '@kmlId',
                                    '@subtramo',
                                    '@tramo'])
lector.start(ruta_kmz)
id_objeto = 528600
fo_net = '31'
infra_net = '30'
nombres_cables = ['COBOS - EL BRACHO TRANSENER','EL BRACHO - COBOS']
patrones_estilo = ['inline', 'msn_ylw-pushpin', '#route', 'geocode']
diccionario_sitios = {}

def poblarSitios():
    wb = load_workbook(filename=path_excel_sitios)
    hoja = wb.active
    for i in range(2,hoja.max_row + 1):
            acronimo = hoja.cell(row=i, column=1).value
            nombre_sitio = hoja.cell(row=i, column=3).value
            departamento = hoja.cell(row=i, column=5).value
            poblacion = hoja.cell(row=i, column=6).value
            proyecto = hoja.cell(row=i, column=7).value
            latitud = hoja.cell(row=i, column=8).value
            longitud = hoja.cell(row=i, column=9).value
            at = hoja.cell(row=i, column=10).value
            convenio  = hoja.cell(row=i, column=11).value
            fecha_aprobada = hoja.cell(row=i, column=12).value
            row_data = {
                "@oName": acronimo,
                "@nombre": nombre_sitio,
                "@departamento": departamento,
                "@pob": poblacion,
                "@proyecto": proyecto,
                "@lat": latitud,
                "@lon": longitud,
                "@at": at,
                "@convenio": convenio,
                "@fecha": fecha_aprobada
            }

            # Add the data to the dictionary using nombre_sitio as the key
            diccionario_sitios[nombre_sitio] = row_data
poblarSitios()

for i in lector.dict_objetos.values():
    nombre = i['Nombre']
    detalle = i['Descripción']
    estilo = i['Estilo']
    coordenadas = i['Coordenadas']
    val = {}



    if any(patron in estilo for patron in patrones_estilo) or nombre in nombres_cables:
        prov = PROVINCIA
    # Cuando es un cable se crea el cable y el tubo en las mismas coordenadas
        tramo = nombre
        subtramo = ''
        if 'DERIVACION' in detalle:
            subtramo = nombre
            tramo = "SIN DATOS"
        migrador.crear_objeto(id=id_objeto,
                            oType='gc/fo',
                            vectores=coordenadas,
                            vals={
                                '@tramo': f'{tramo}',
                                '@subtramo': f'{subtramo}',
                                '@propietarioCable': '1',
                                '@mediciones': 'true',
                                '@prov': f'{prov.capitalize()}',
                                '@capacidad':'48',
                                '@buffers':'2',
                                '@marcaFO':'9',
                                "@estado":"E",
                                '@propietarioInfra':'1',
                                '@datosFO':'5',
                                '@ductosO':'2',
                                '@tipoRed':'1',
                                '@metodo':'1',
                            },
                            nID=fo_net)
        id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración
        migrador.crear_objeto(id=id_objeto,
                            oType='gc/duc',
                            vectores=coordenadas,
                            vals={
                                '@oName': f'{nombre.upper()}',
                                '@nc1': f'{nombre.upper()}',
                                '@ocupacion1': 'true',
                                '@estado': 'E',
                                '@tendido': '1',
                                '@propietarioTri': '1',
                            },
                            nID= infra_net)
        id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración

    elif '#m_ylw-pushpin' in estilo:
        migrador.crear_objeto(id=id_objeto,
                            oType='go/fo/pos',
                            vectores=coordenadas,
                            vals={
                                    "@oName": f'{nombre}',
                                    "@material": "1",
                                    "@EstadoPo": "0",
                                    "@estado":"E",
                                    "@tipoPo":"T",
                                    "@propietarioPos":"1",
                                },
                            nID= infra_net)
        id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración

    elif '#msn_icon' in estilo or '#msn_flag' in estilo or 'square' in estilo:
        #Si dice la palabra BOX en el nombre se crea un empalme y una caja
        if 'BOX' in nombre:
            migrador.crear_objeto(id=id_objeto,
                            oType='go/fo/cam',
                            vectores=coordenadas,
                            vals={
                                    "@oName": f'{detalle}',
                                    "@tapas": "4",
                                    "@tipoCam": "1",
                                    "@estado":"E",
                                    "@estadoCam":"0",
                                    "@aCaja": f"{nombre}"
                                },
                            nID= infra_net)
            id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración
            migrador.crear_objeto(  id=id_objeto,
                                    oType='go/fo/em',
                                    vectores=coordenadas,
                                    vals={
                                        "@oName": f'{nombre}',
                                        "@posicion": "1",
                                        "@tipoCaja": "5",
                                        "@tipoEmpalme":"2",
                                        "@estadoEm":"E",
                                        "@ocupacion":"1"
                                        },
                                    nID=fo_net)
            id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración
        #Si no lo dice se crea una caja unicamente
        else:
            migrador.crear_objeto(id=id_objeto,
                                oType='go/fo/cam',
                                vectores=coordenadas,
                                vals={
                                    "@oName":f'{nombre}',
                                    "@tapas":"3",
                                    "@tipoCam":"3",
                                    "@estado":"E",
                                    "@propietarioCam":"1",
                                    "@estadoCam":"0"
                                },
                                nID= infra_net)
            id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración

    elif 'square__' in estilo or '#msn_grn-stars' in estilo:
        print(estilo)
        print(nombre)
        #se crea una caja y un empalme, pero es otro tipo de icono en el kmz
        migrador.crear_objeto(id=id_objeto,
                            oType='go/fo/cam',
                            vectores=coordenadas,
                            vals={
                                    "@oName": f'{detalle}',
                                    "@tapas": "4",
                                    "@tipoCam": "1",
                                    "@estado":"E",
                                    "@estadoCam":"0",
                                    "@aCaja": f"{nombre}"
                                },
                            nID= infra_net)
        id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración
        migrador.crear_objeto(  id=id_objeto,
                                oType='go/fo/em',
                                vectores=coordenadas,
                                vals={
                                    "@oName": f'{nombre}',
                                    "@posicion": "1",
                                    "@tipoCaja": "5",
                                    "@tipoEmpalme":"2",
                                    "@estadoEm":"E",
                                    "@ocupacion":"1"
                                    },
                                nID=fo_net)
        id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración

    elif 'placemark_circle' in estilo:
        prov = PROVINCIA
        if nombre in diccionario_sitios:
            datos_sitio = diccionario_sitios[nombre]
            # Acceder a información específica del sitio
            acronimo = datos_sitio["@oName"]
            if 'FALSO' in acronimo:
                acronimo = ''
            departamento = datos_sitio["@departamento"]
            poblacion = datos_sitio["@pob"]
            proyecto = datos_sitio["@proyecto"]
            latitud = datos_sitio["@lat"]
            longitud = datos_sitio["@lon"]
            at = datos_sitio["@at"]
            if at:
                at = at.strftime('%Y-%m-%d')
            convenio = datos_sitio["@convenio"]
            if convenio:
                convenio = convenio.strftime('%Y-%m-%d')
            fecha_aprobada = datos_sitio["@fecha"]
            if fecha_aprobada:
                fecha_aprobada = fecha_aprobada.strftime('%Y-%m-%d')
            migrador.crear_objeto(id=id_objeto,
                            oType='go/fo/si',
                            vectores=coordenadas,
                            vals={
                                '@nombre': f'{nombre.upper()}',
                                '@oName': f'{acronimo}',
                                '@prov': f'{prov.capitalize()}',
                                '@estado': 'E',
                                '@tipo':'2',
                                '@proyecto':proyecto,
                                '@lat':latitud,
                                '@lon':longitud,
                                '@at':at,
                                '@convenio':convenio,
                                '@fecha':fecha_aprobada,
                            },
                            nID= infra_net)
            id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración
            migrador.crear_objeto(id=id_objeto,
                            oType='go/fo/gb',
                            vectores=coordenadas,
                            vals={
                                '@oName': f'{nombre.upper()}',
                                '@acronimo': f'R{acronimo}',
                                '@prov': f'{prov.capitalize()}',
                                '@estado': 'E',
                            },
                            nID= infra_net)
            id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración
        else:
            migrador.crear_objeto(id=id_objeto,
                                oType='go/fo/si',
                                vectores=coordenadas,
                                vals={
                                    '@nombre': f'{nombre.upper()}',
                                    '@prov': f'{prov.capitalize()}',
                                    '@estado': 'E',
                                    '@tipo':'2',
                                },
                                nID= infra_net)
            id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración
    else:
        print(f'Nombre: {nombre}')
        print(f'Estilo: {estilo}')
        print('------------------')