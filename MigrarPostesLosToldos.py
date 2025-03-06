from KMZReader.db_migration import MigrarInfo
from openpyxl import load_workbook
import re

# Suponiendo que ya tienes el contenido_kml del archivo KML
ruta_excel_cables = 'C:/Users/j2sae/Documents/Postes.xlsx'
COMPANY_ID = '132'
# En el objeto MigrarInfo se asignan las variables del ID de la compania y las variables que seran cargadas tambien en el sidx
migrador = MigrarInfo(company_id='132',
                    variables_sidx=['@oName',
                                    '@kmlId',
                                    ])

objectID = 7500
fo_net = '1'
infra_net = '4'
clientNetID = '2'
localidad_lista= []

wb = load_workbook(filename=ruta_excel_cables)
hoja = wb.active
unixtime = "1..1."
for i in range(5,hoja.max_row + 1):

    nombre = hoja.cell(row=i, column=1).value
    altura = hoja.cell(row=i, column=3).value
    tipo = hoja.cell(row=i, column=4).value
    latitud = hoja.cell(row=i, column=8).value
    longitud = hoja.cell(row=i, column=9).value
    latitudConvertida = 
    longitudConvertida = 
    coordenadas = [[latitud,longitud]]
    migrador.crear_objeto(id=objectID,
                          oType='go/pos',
                          vectores=coordenadas,
                          vals={
                              '@foType': '12'
                                },
                          nID=fo_net)
    objectID += 1  # Incrementar el ID del objeto para la siguiente iteración