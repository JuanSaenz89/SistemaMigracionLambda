
from KMZReader.db_migration import MigrarInfo
from openpyxl import load_workbook

# Suponiendo que ya tienes el contenido_kml del archivo KML
ruta_excel = '/home/linux/Downloads/Red por Cliente.xlsx'

# En el objeto MigrarInfo se asignan las variables del ID de la compania y las variables que seran cargadas tambien en el sidx
migrador = MigrarInfo(company_id='131',
                    variables_sidx=['@oName',
                                    '@kmlId',
                                    ])

id_objeto = 2413
fo_net = '1'
infra_net = '4'
cliente_net = '2'
localidad_lista= []
    
wb = load_workbook(filename=ruta_excel)
hoja = wb['Red Por Cliente']
unixtime = "1..1."
for i in range(5,hoja.max_row + 1):
    numero = hoja.cell(row=i, column=1).value
    nombre = hoja.cell(row=i, column=2).value
    domicilio = hoja.cell(row=i, column=3).value
    latitud = hoja.cell(row=i, column=4).value
    longitud = hoja.cell(row=i, column=5).value
    plan = hoja.cell(row=i, column=6).value
    bloqueo = hoja.cell(row=i, column=7).value
    usuario = hoja.cell(row=i, column=8).value
    serialNumber = hoja.cell(row=i, column=9).value
    ontId = hoja.cell(row=i, column=10).value
    ip = hoja.cell(row=i, column=11).value
    mac = hoja.cell(row=i, column=12).value
    cableModem = hoja.cell(row=i, column=13).value
    ssid = hoja.cell(row=i, column=14).value
    claveSsid = hoja.cell(row=i, column=15).value
    ssid5 = hoja.cell(row=i, column=16).value
    claveSsid5 = hoja.cell(row=i, column=17).value
    padreDirecto = hoja.cell(row=i, column=19).value
    padre = hoja.cell(row=i, column=19).value
    com = hoja.cell(row=i, column=20).value

    if not latitud or not longitud:
        continue
    coordenadas = [[[longitud , latitud]]]

    vals = {
        '@oName': numero,
        '@nombre': nombre,
        '@domicilio': domicilio,
        '@plan': plan,
        '@bloqueo': bloqueo,
        '@usuario': usuario,
        '@serialNumber': serialNumber,
        '@ontId': ontId,
        '@ip': ip,
        '@mac': mac,
        '@cableModem': cableModem,
        '@ssid': ssid,
        '@claveSsid': claveSsid,
        '@ssid5': ssid5,
        '@claveSsid5': claveSsid5,
        '@padreDirecto': padreDirecto,
        '@padre': padre,
        '@com': com,
    }
    # Filtrar las claves con valores no None
    vals_filtrados = {k: v for k, v in vals.items() if v is not None}

    migrador.crear_objeto(id=id_objeto,
                          oType='go/fo/cli',
                          vectores=coordenadas,
                          vals=vals_filtrados,
                          nID=cliente_net)
    id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración

localidad_lista = set(localidad_lista)
print(localidad_lista)