
from KMZReader.db_migration import MigrarInfo
from openpyxl import load_workbook
import time
import re


COMPANY_ID = '132'
UNIXTIME = int(time.time())

class MigradorBCM():

    # En el objeto MigrarInfo se asignan las variables del ID de la compania y las variables que seran cargadas tambien en el sidx
    migrador = MigrarInfo(company_id='132',
                    variables_sidx=['@oName',
                                    '@kmlId',
                                    ])
    
    def __init__(self):

        # Aqui van las rutas de los excels de clientes y de Splitters y Naps
        self.ruta_excel = 'C:/Users/j2sae/Documents/Listado por Origen - Los Toldos.xlsx'
        self.ruta_excel_clientes = 'C:/Users/j2sae/Documents/Clientes Los Toldos.xlsx'

        self.objectID = 1 # ID en el que se comienza a crear los objetos
        self.fo_net = 1#input('Ingrese el Id de la capa de FO: ')
        self.infra_net = 4#input('Ingrese el Id de la capa de Infraestructura: ')
        self.clientNetID = 2#input('Ingrese el Id de la capa de Clientes: ')

        self.nameList = []
        self.CierresDict = {}
        self.EmpalmesDict = {}
        self.ClientsDict = {}
        self.CierresOrdenadosPorPadreDict = {}
    

    def start(self):
        self.migrateCierres(self.ruta_excel)
        self.migrateClientes(self.ruta_excel_clientes)
        self.conectarClientesANaps()
        #self.conectarNaps()

    def migrateCierres(self, filename):
        wb = load_workbook(filename = filename)
        hoja = wb['Listado por Origen']

        for i in range(5,hoja.max_row + 1):
            id = hoja.cell(row=i, column=1).value
            origen = hoja.cell(row=i, column=2).value
            equipamiento = hoja.cell(row=i, column=6).value
            nombre = hoja.cell(row=i, column=6).value
            latitud = hoja.cell(row=i, column=8).value
            longitud = hoja.cell(row=i, column=9).value
            tipo = hoja.cell(row=i, column=10).value
            padreDirecto = hoja.cell(row=i, column=11).value
            direccion = hoja.cell(row=i, column=16).value

            if not latitud or not longitud:
                continue
            
            coordenadas = [[[longitud , latitud]]]
            
            if tipo == 'FO: Distribución / NAP':

                if '_' in equipamiento:
                    equipamiento = equipamiento.split('_')
                    nombre = equipamiento[1]
                
                vals = {
                    '@oName': nombre,
                    '@oType': '2N',
                    '@ref': direccion
                }

                self.migrador.crear_objeto(id=self.objectID,
                                oType='go/fo/cie',
                                vectores=coordenadas,
                                vals=vals,
                                nID=self.fo_net)
                
                self.poblarDiccionarioCierres(nombre, padreDirecto, coordenadas, self.CierresDict)

                self.objectID += 1  # Incrementar el ID del objeto para la siguiente iteración

            elif tipo == 'FO: Sangrado / FDH':

                if '_' in equipamiento:
                    equipamiento = equipamiento.split('_')
                    nombre = equipamiento[1]

                if nombre not in self.nameList:
                    self.nameList.append(nombre)

                elif nombre in self.nameList:
                    print(nombre)
                    continue

                vals = {
                    '@oName': nombre,
                    '@oType': '1N',
                    '@ref': direccion
                }
                self.migrador.crear_objeto(id=self.objectID,
                                oType='go/fo/cie',
                                vectores=coordenadas,
                                vals=vals,
                                nID=self.fo_net)
                
                self.poblarDiccionarioCierres(nombre, padreDirecto, coordenadas, self.EmpalmesDict)

                self.objectID += 1 

    def poblarDiccionarioCierres(self, nombre, padreDirecto, coordenadas, diccionario):

        diccionario[self.objectID] = {
                                        'Vectores': coordenadas,
                                        'Nombre': nombre,
                                        'Padre': padreDirecto
                                        }
        if diccionario == self.CierresDict:
                self.CierresOrdenadosPorPadreDict.setdefault(padreDirecto, []).append({
                'Vectores': coordenadas,
                'Nombre': nombre,
                'ID': self.objectID
                })
    
    def poblarDiccionarioClientes(self, padreDirecto, coordenadas):
        self.ClientsDict[self.objectID] = {
                                            'Vectores' : coordenadas,
                                            'Padre' : padreDirecto
                                            }

    def migrateClientes(self, filename):
        wb = load_workbook(filename=filename)
        hoja = wb['Coordenada Parceada']
        unixtime = "1..1."

        for i in range(2,hoja.max_row + 1):
            numero = hoja.cell(row=i, column=1).value
            nombre = hoja.cell(row=i, column=2).value
            apellido = hoja.cell(row=i, column=3).value
            if nombre and apellido:
                nombreCompleto = nombre + ' ' + apellido
            elif apellido:
                nombreCompleto = apellido
            elif nombre:
                nombreCompleto = nombre
            domicilio = hoja.cell(row=i, column=4).value
            latitud = hoja.cell(row=i, column=19).value
            longitud = hoja.cell(row=i, column=20).value
            plan = hoja.cell(row=i, column=11).value
            bloqueo = hoja.cell(row=i, column=13).value
            usuario = hoja.cell(row=i, column=15).value
            serialNumber = hoja.cell(row=i, column=30).value
            ontId = hoja.cell(row=i, column=31).value
            ip = hoja.cell(row=i, column=16).value
            mac = hoja.cell(row=i, column=17).value
            cableModem = hoja.cell(row=i, column=18).value
            ssid = hoja.cell(row=i, column=26).value
            claveSsid = hoja.cell(row=i, column=27).value
            ssid5 = hoja.cell(row=i, column=28).value
            claveSsid5 = hoja.cell(row=i, column=29).value
            padreDirecto = hoja.cell(row=i, column=22).value
            com = hoja.cell(row=i, column=10).value

            if not latitud or not longitud:
                continue

            coordenadas = [[[longitud , latitud]]]

            vals = {
                '@oName': str(numero),
                '@nombre': nombreCompleto,
                '@domicilio': str(domicilio),
                '@plan': plan,
                '@bloqueo': bloqueo,
                '@usuario': usuario,
                '@serialNumber': serialNumber,
                '@ontId': str(ontId),
                '@ip': ip,
                '@mac': mac,
                '@cableModem': cableModem,
                '@ssid': ssid,
                '@claveSsid': claveSsid,
                '@ssid5': ssid5,
                '@claveSsid5': claveSsid5,
                '@padreDirecto': padreDirecto,
                '@com': com,
            }
            # Filtrar las claves con valores no None
            vals_filtrados = {k: v for k, v in vals.items() if v is not None}
            
            self.migrador.crear_objeto(id=self.objectID,
                                oType='go/fo/cli',
                                vectores=coordenadas,
                                vals=vals_filtrados,
                                nID=self.clientNetID)
            
            self.poblarDiccionarioClientes(padreDirecto, coordenadas)

            self.objectID += 1 # Incrementar el ID del objeto para la siguiente iteración    

    def conectarClientesANaps(self):
        for oID, dictCliente in self.ClientsDict.items():
            padre = dictCliente['Padre']
            padreArreglado = self.cambiarNombrePadre(padre)
            vectorCliente = dictCliente['Vectores']

            for cierreOID, dictCierre in self.CierresDict.items():
                nombreCierre = dictCierre['Nombre']
                vectorCierre = dictCierre['Vectores']

                if padre == nombreCierre or padreArreglado == nombreCierre:
                    self.crearCable(oID, vectorCliente, cierreOID, vectorCierre,'DROP')
                    self.objectID += 1

    def crearCable(self, idUno, coordenadaUno, idDos, coordenadaDos, tipoCable):
        idCable = f"{COMPANY_ID}.{self.fo_net}.{self.objectID}"
        idCliente = f"{COMPANY_ID}.{self.clientNetID}.{idUno}"
        idCierre = f"{COMPANY_ID}.{self.fo_net}.{idDos}"
        longitudUno = coordenadaUno[0][0][0]  # Primer elemento de la primera lista
        latitudUno = coordenadaUno[0][0][1]
        longitudDos = coordenadaDos[0][0][0]  # Primer elemento de la primera lista
        latitudDos = coordenadaDos[0][0][1]

        with open('CABLES.txt','a') as ffcables:
            # escribimos el O
            ffcables.write(f'SET {idCable} "{UNIXTIME}..1."\n') 
            # escribimos el OCFG
            ffcables.write(f'SADD {idCable}:ocfg "{UNIXTIME}..1.:gc/fo"\n') 
            # escribimos los VALS
            ffcables.write(f'SADD {idCable}:val "{UNIXTIME}..1.:@foType|{tipoCable}|0"\n')
            # escribimos los V (Vectores) 
            ffcables.write(f'SADD {idCable}:v "{UNIXTIME}..1.:{latitudUno}|{longitudUno}|0|0"\n')
            ffcables.write(f'SADD {idCable}:v "{UNIXTIME}..1.:{latitudDos}|{longitudDos}|1|0"\n') 
            # escribimos los GEO
            ffcables.write(f'GEOADD {COMPANY_ID}.{self.fo_net}:geoidx {longitudUno} {latitudUno} "{UNIXTIME}..1.:{idCable}|0|2"\n')
            ffcables.write(f'GEOADD {COMPANY_ID}.{self.fo_net}:geoidx {longitudDos} {latitudDos} "{UNIXTIME}..1.:{idCable}|1|2"\n')
            # hacemos las conexiones
            ffcables.write(f'SADD {idCliente}:co "{UNIXTIME}..1.:2|{idCable}|1"\n')
            ffcables.write(f'SADD {idCable}:co "{UNIXTIME}..1.:1|{idCliente}|2"\n')
            ffcables.write(f'SADD {idCierre}:co "{UNIXTIME}..1.:1|{idCable}|2"\n')
            ffcables.write(f'SADD {idCable}:co "{UNIXTIME}..1.:2|{idCierre}|1"\n')
            # Creamos los objetos internos

        self.objectID += 1
        idObjetoInterno = f"{COMPANY_ID}.100.{self.objectID}"
        
        vals = {
                '@color' : 'GR',
                '@io' : f'{idCable}',
                '@io0' : f'{idCable}',
                '@name' : '1',
                '@order' : '0'
                }
        
        self.crearIO(idCable, idObjetoInterno, 'io/fo/t', vals)

    def crearIO(self, Padre, ID, oType, vals):
        with open('CABLES.txt','a') as ffcables:
            ffcables.write(f'SADD {Padre}:io "{UNIXTIME}..1.:{ID}"\n')
            ffcables.write(f'SET {ID} "{UNIXTIME}..1."\n') 
            # escribimos el OCFG
            ffcables.write(f'SADD {ID}:ocfg "{UNIXTIME}..1.:{oType}"\n') 
            # escribimos los VALS
            for key, value in vals.items():
                ffcables.write(f'SADD {ID}:val "{UNIXTIME}..1.:{key}|{value}|0"\n')
        
        if oType == 'io/fo/t':
            self.objectID += 1
            idObjetoInterno = f"{COMPANY_ID}.100.{self.objectID}"
            vals = {
                    '@color' : 'GR',
                    '@io' : f'{ID}',
                    '@io0' : f'{ID}',
                    '@name' : '1',
                    '@order' : '0'
                    }
            self.crearIO(ID, idObjetoInterno, 'io/fo/h', vals)

    def cambiarNombrePadre(self, padre):
        try:
            pattern = r"(Caja\d+-Nodo\d+)"
            match = re.search(pattern, padre)
            return match.group(1) if match else None
        except TypeError:
            return ''

    def conectarNaps(self):
        count = 0
        for padre, naps in self.CierresOrdenadosPorPadreDict.items():
            cantidadCajas = len(naps)

            for i in range(cantidadCajas - 1):

                caja_actual = naps[i]
                caja_siguiente = naps[i + 1]
                idNapUno = caja_actual['ID']
                idNapDos = caja_siguiente['ID']
                vectorNapUno = caja_actual['Vectores']
                vectorNapDos = caja_siguiente['Vectores']
                self.crearCable(idNapUno, vectorNapUno, idNapDos, vectorNapDos,'DROP')
                 # Llamamos a la función para crear la conexión


MigradorBCM().start()