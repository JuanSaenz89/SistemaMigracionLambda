
from KMZReader.db_migration import MigrarInfo
from openpyxl import load_workbook
import re
# Suponiendo que ya tienes el contenido_kml del archivo KML
ruta_excel = '/home/linux/Downloads/Listado por Origen - Estandarizado - Lujan.xlsx'

# En el objeto MigrarInfo se asignan las variables del ID de la compania y las variables que seran cargadas tambien en el sidx
migrador = MigrarInfo(company_id='131',
                    variables_sidx=['@oName',
                                    '@kmlId',
                                    ])

id_objeto = 1
fo_net = '1'
infra_net = '4'
cliente_net = '2'
lista_nombres = []
    
wb = load_workbook(filename=ruta_excel)
hoja = wb['Listado por Origen']
unixtime = "1..1."
for i in range(5,hoja.max_row + 1):
    id = hoja.cell(row=i, column=1).value
    origen = hoja.cell(row=i, column=2).value
    equipamiento = hoja.cell(row=i, column=6).value
    nombre = hoja.cell(row=i, column=6).value
    latitud = hoja.cell(row=i, column=8).value
    longitud = hoja.cell(row=i, column=9).value
    tipo = hoja.cell(row=i, column=10).value
    padreDirecto = hoja.cell(row=i, column=11).value
    direccion = hoja.cell(row=i, column=16).value

    if not latitud or not longitud:
        continue
    coordenadas = [[[longitud , latitud]]]
    
    if tipo == 'FO: Distribución / NAP':
        if '_' in equipamiento:
            equipamiento = equipamiento.split('_')
            nombre = equipamiento[1]
        vals = {
            '@oName': nombre,
            '@oType': '2N',
            '@ref': direccion
        }
        migrador.crear_objeto(id=id_objeto,
                          oType='go/fo/cie',
                          vectores=coordenadas,
                          vals=vals,
                          nID=fo_net)
        id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración
    elif tipo == 'FO: Sangrado / FDH':

        if '_' in equipamiento:
            equipamiento = equipamiento.split('_')
            nombre = equipamiento[1]

        if nombre not in lista_nombres:
            lista_nombres.append(nombre)

        elif nombre in lista_nombres:
            print(nombre)
            continue

        vals = {
            '@oName': nombre,
            '@oType': '1N',
            '@ref': direccion
        }
        migrador.crear_objeto(id=id_objeto,
                          oType='go/fo/cie',
                          vectores=coordenadas,
                          vals=vals,
                          nID=fo_net)
        id_objeto += 1  # Incrementar el ID del objeto para la siguiente iteración
    
    
