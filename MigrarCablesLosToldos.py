from KMZReader.db_migration import MigrarInfo
from openpyxl import load_workbook
import re

# Suponiendo que ya tienes el contenido_kml del archivo KML
ruta_excel_cables = 'C:/Users/j2sae/Documents/Cables - Los Toldos.xlsx'
COMPANY_ID = '132'
# En el objeto MigrarInfo se asignan las variables del ID de la compania y las variables que seran cargadas tambien en el sidx
migrador = MigrarInfo(company_id='132',
                    variables_sidx=['@oName',
                                    '@kmlId',
                                    ])

objectID = 2413
fo_net = '1'
infra_net = '4'
clientNetID = '2'
localidad_lista= []

wb = load_workbook(filename=ruta_excel_cables)
hoja = wb.active
unixtime = "1..1."
for i in range(5,hoja.max_row + 1):

    ladoUno = hoja.cell(row=i, column=1).value
    ladoDos = hoja.cell(row=i, column=1).value
    vectores = hoja.cell(row=i, column=2).value #TODO limpiar vectores

    patron = r"(-?\d+\.\d+)@(-?\d+\.\d+)"
    coordenadas = re.findall(patron, vectores)
    coordenadas = []

    migrador.crear_objeto(id=objectID,
                          oType='gc/fo',
                          vectores=coordenadas,
                          vals={
                              '@oName': '',
                              '@foType': ''
                                },
                          nID=fo_net)
    objectID += 1  # Incrementar el ID del objeto para la siguiente iteración